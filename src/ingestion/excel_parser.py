"""
excel_parser.py
===============
.xlsx / .xlsm 파일을 시트 단위로 파싱한다.

특징:
- openpyxl 사용 (data_only=True 로 수식 결과값을 읽는다).
- 시트별로 raw_table_text 를 만들어 ParsedSection 으로 반환.
- 너무 긴 시트는 N행 단위 블록으로 나누어 여러 ParsedSection 으로 반환.
- content_type 은 "excel_raw_table".
- 병합 셀이 있을 경우, 병합 영역 좌상단 값을 모든 셀에 채워준다.
- 이미지/OCR 은 TODO.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from src.logger import get_logger

log = get_logger(__name__)

# 시트 1개에서 너무 큰 표를 그대로 두면 컨텍스트가 폭발하므로 블록 단위로 끊어준다.
DEFAULT_BLOCK_ROWS = 60
MAX_CELL_LEN = 200  # 셀 1개의 최대 길이 (긴 셀은 잘라낸다)


def _coerce_cell(value: Any) -> str:
    """셀 값을 문자열로 정규화."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    s = str(value).strip()
    if len(s) > MAX_CELL_LEN:
        s = s[: MAX_CELL_LEN - 3] + "..."
    return s


def _build_merged_lookup(ws) -> Dict[Tuple[int, int], Any]:
    """병합 셀 좌표 → 좌상단 값 매핑."""
    lookup: Dict[Tuple[int, int], Any] = {}
    try:
        for rng in list(ws.merged_cells.ranges):
            min_row, min_col = rng.min_row, rng.min_col
            max_row, max_col = rng.max_row, rng.max_col
            top_left = ws.cell(row=min_row, column=min_col).value
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    lookup[(r, c)] = top_left
    except Exception:  # noqa: BLE001
        pass
    return lookup


def _format_row(row: List[str]) -> str:
    """행 1줄을 탭으로 이어 붙이는데, 모두 빈 칸이면 빈 문자열 반환."""
    if not any(cell for cell in row):
        return ""
    return "\t".join(row)


def parse_excel(path: Path, document_id: str) -> List[Dict[str, Any]]:
    """
    Excel 파일을 시트 단위로 파싱한다.

    Returns
    -------
    list of dict
        각 dict 는 ParsedSection 형태와 호환:
          - section_title
          - content_type ("excel_raw_table")
          - content (raw table text)
          - metadata (sheet_name, max_row, max_column, block_index, ...)
    """
    try:
        from openpyxl import load_workbook  # 지연 import (의존성 부담 감소)
    except ImportError as e:
        raise RuntimeError(
            "openpyxl 이 설치되어 있지 않습니다. requirements.txt 를 다시 설치하세요."
        ) from e

    sections: List[Dict[str, Any]] = []
    try:
        wb = load_workbook(filename=str(path), data_only=True, read_only=True)
    except Exception as e:  # 파일 손상/포맷 이상
        log.error("Excel load 실패: %s (%s)", path.name, e)
        raise

    try:
        for sheet_name in wb.sheetnames:
            try:
                ws = wb[sheet_name]
            except Exception as e:  # noqa: BLE001
                log.warning("시트 로드 실패: %s/%s (%s)", path.name, sheet_name, e)
                continue

            # read_only 모드에서는 merged_cells 가 비어있을 수 있음
            merged_lookup = _build_merged_lookup(ws) if not ws.__class__.__name__.startswith("ReadOnly") else {}

            rows_text: List[str] = []
            max_col = 0
            row_count = 0
            for row in ws.iter_rows(values_only=False):
                cells: List[str] = []
                for cell in row:
                    val = cell.value
                    if val is None and (cell.row, cell.column) in merged_lookup:
                        val = merged_lookup[(cell.row, cell.column)]
                    cells.append(_coerce_cell(val))
                # 우측 빈 칸 제거
                while cells and not cells[-1]:
                    cells.pop()
                if not cells:
                    continue
                line = _format_row(cells)
                if line:
                    rows_text.append(line)
                    max_col = max(max_col, len(cells))
                row_count += 1

            if not rows_text:
                log.info("빈 시트 skip: %s/%s", path.name, sheet_name)
                continue

            # 너무 큰 시트는 N행 단위 블록으로 나눈다
            for block_index, start in enumerate(range(0, len(rows_text), DEFAULT_BLOCK_ROWS)):
                block_lines = rows_text[start : start + DEFAULT_BLOCK_ROWS]
                block_text = "\n".join(block_lines)
                section_title = f"{sheet_name}" if block_index == 0 else f"{sheet_name} (블록 {block_index + 1})"
                sections.append({
                    "section_title": section_title,
                    "content_type": "excel_raw_table",
                    "content": block_text,
                    "metadata": {
                        "sheet_name": sheet_name,
                        "block_index": block_index,
                        "block_row_count": len(block_lines),
                        "sheet_row_count": len(rows_text),
                        "sheet_col_count": max_col,
                        "table_range": f"{sheet_name}!1:{len(rows_text)}",
                    },
                })

    finally:
        try:
            wb.close()
        except Exception:  # noqa: BLE001
            pass

    log.info("Excel 파싱 완료: %s -> %d section", path.name, len(sections))
    return sections
